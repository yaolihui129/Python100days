#!/usr/bin/env python
# -*- coding: utf-8 -*-
# 读取Excel文件
# @Date    : 2019-07-18 22:32:51
# @Author  : yaolh (yaolihui129@sina.com)
# @Link    : https://github.com/yaolihui129
# @Version : 0.1

from openpyxl import load_workbook
from openpyxl import Workbook

workbook = load_workbook('./全班学生数据.xlsx')
print(workbook.sheetnames)
sheet = workbook[workbook.sheetnames[0]]
print(sheet.title)
for row in range(2,7):
	for col in range(2,7):
		cell_index = chr(col) + str(row)
		print(sheet[cell_index].value,end='\t')
	print()